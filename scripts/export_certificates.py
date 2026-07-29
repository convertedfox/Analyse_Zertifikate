from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import TypedDict

import openpyxl

EXPECTED_HEADERS = ("Typ", "Zertifikat", "Modulnummer", "Modulname", "REF")


@dataclass(frozen=True)
class ModuleRecord:
    module_id: str
    module_name: str


class SourceInfo(TypedDict):
    excel_path: str
    sheet_main: str
    sheet_selected: str


class ModulePayload(TypedDict):
    module_id: str
    module_name: str


class CertificatePayload(TypedDict):
    certificate_name: str
    modules_by_type: dict[str, list[ModulePayload]]


class DatasetPayload(TypedDict):
    source: SourceInfo
    selected_certificate_names: list[str]
    certificates: list[CertificatePayload]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export certificate modules to JSON")
    parser.add_argument(
        "--excel",
        type=Path,
        default=Path("data/Analyse der Zertifikate.xlsx"),
        help="Path to source Excel file",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("data/certificates.json"),
        help="Path to generated JSON file",
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="Only validate source and print summary without writing JSON",
    )
    return parser.parse_args()


def validate_headers(sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
    headers = tuple(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[:5])
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected headers: {headers}")


def normalize(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_dataset(excel_path: Path) -> DatasetPayload:
    workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    source_sheet = workbook["Zertifikate und deren Module"]
    selected_sheet = workbook["Einzelne Zertifikate"]

    validate_headers(source_sheet)

    modules_by_name_and_type: dict[str, dict[str, dict[str, ModuleRecord]]] = defaultdict(
        lambda: defaultdict(dict)
    )
    module_name_conflicts: dict[str, set[str]] = defaultdict(set)

    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        cert_type = normalize(row[0])
        cert_name = normalize(row[1])
        module_id = normalize(row[2])
        module_name = normalize(row[3])

        if not cert_type and not cert_name and not module_id and not module_name:
            continue

        if cert_type not in {"CAS", "DAS"}:
            raise ValueError(f"Unexpected certificate type: {cert_type!r}")
        if not cert_name or not module_id or not module_name:
            raise ValueError(f"Incomplete row for {cert_name!r} ({cert_type}): {row[:4]!r}")

        existing = modules_by_name_and_type[cert_name][cert_type].get(module_id)
        if existing and existing.module_name != module_name:
            raise ValueError(
                f"Conflicting module names for {module_id!r}: "
                f"{existing.module_name!r} vs {module_name!r}"
            )

        modules_by_name_and_type[cert_name][cert_type][module_id] = ModuleRecord(
            module_id=module_id,
            module_name=module_name,
        )
        module_name_conflicts[module_id].add(module_name)

    for module_id, names in module_name_conflicts.items():
        if len(names) > 1:
            raise ValueError(f"Module {module_id!r} has multiple names: {sorted(names)!r}")

    selected_names: list[str] = []
    for row in selected_sheet.iter_rows(min_row=2, values_only=True):
        name = normalize(row[0])
        if name:
            selected_names.append(name)

    if len(set(selected_names)) != len(selected_names):
        raise ValueError("Sheet 'Einzelne Zertifikate' contains duplicates")

    unknown = [name for name in selected_names if name not in modules_by_name_and_type]
    if unknown:
        raise ValueError(f"Selected certificates missing in source sheet: {unknown!r}")

    certificates: list[CertificatePayload] = []
    for cert_name in selected_names:
        modules_by_type: dict[str, list[ModulePayload]] = {}
        for cert_type in ("CAS", "DAS"):
            modules = modules_by_name_and_type[cert_name].get(cert_type, {})
            modules_by_type[cert_type] = [
                {"module_id": module.module_id, "module_name": module.module_name}
                for module in sorted(modules.values(), key=lambda item: item.module_id)
            ]

        certificates.append(
            {
                "certificate_name": cert_name,
                "modules_by_type": modules_by_type,
            }
        )

    return {
        "source": {
            "excel_path": str(excel_path.as_posix()),
            "sheet_main": "Zertifikate und deren Module",
            "sheet_selected": "Einzelne Zertifikate",
        },
        "selected_certificate_names": selected_names,
        "certificates": certificates,
    }


def write_json(payload: DatasetPayload, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    args = parse_args()
    payload = build_dataset(args.excel)

    if args.check:
        print(
            f"Validated {len(payload['selected_certificate_names'])} certificates from {args.excel}"
        )
        return

    write_json(payload, args.output)
    print(f"Wrote {args.output} with {len(payload['selected_certificate_names'])} certificates")


if __name__ == "__main__":
    main()
